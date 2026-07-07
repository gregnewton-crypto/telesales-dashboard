#!/usr/bin/env python3
"""Generate Excel workbook and Apps Script from airtable_8_tables_schema.json."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "airtable_8_tables_schema.json"
XLSX_PATH = ROOT / "airtable_8_tables_schema.xlsx"
APPS_SCRIPT_PATH = ROOT / "sync" / "apps-script" / "SchemaExport.gs"

HEADER_FILL = PatternFill("solid", fgColor="522A10")
HEADER_FONT = Font(bold=True, color="FFE180")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_schema():
    with SCHEMA_PATH.open() as f:
        return json.load(f)


def format_field_options(field):
    ftype = field.get("type", "")
    opts = field.get("options", {})
    parts = []

    if ftype == "number":
        parts.append(f"precision={opts.get('precision', 0)}")
    elif ftype == "date":
        df = opts.get("dateFormat", {})
        parts.append(f"date: {df.get('name')} ({df.get('format')})")
    elif ftype == "dateTime":
        df = opts.get("dateFormat", {})
        tf = opts.get("timeFormat", {})
        parts.append(
            f"date: {df.get('name')} ({df.get('format')}); "
            f"time: {tf.get('name')} ({tf.get('format')}); "
            f"tz: {opts.get('timeZone')}"
        )
    elif ftype == "createdTime":
        result = opts.get("result", {}).get("options", {})
        df = result.get("dateFormat", {})
        tf = result.get("timeFormat", {})
        parts.append(
            f"auto timestamp; date: {df.get('name')} ({df.get('format')}); "
            f"time: {tf.get('name')} ({tf.get('format')}); "
            f"tz: {result.get('timeZone')}"
        )
    elif ftype in ("singleSelect", "multipleSelects"):
        parts.append(f"{len(opts.get('choices', []))} choices (see Select Choices tab)")
    elif ftype == "multipleRecordLinks":
        parts.append(
            f"linkedTableId={opts.get('linkedTableId')}; "
            f"inverseLinkFieldId={opts.get('inverseLinkFieldId')}"
        )
    elif ftype == "multipleLookupValues":
        parts.append(
            f"via {opts.get('recordLinkFieldId')} -> {opts.get('fieldIdInLinkedTable')}; "
            f"result={opts.get('result', {}).get('type')}"
        )
    elif ftype == "rollup":
        parts.append(
            f"via {opts.get('recordLinkFieldId')} -> {opts.get('fieldIdInLinkedTable')}; "
            f"result={opts.get('result', {}).get('type')}"
        )
        if opts.get("formula"):
            parts.append(f"formula={opts['formula']}")
    elif ftype == "formula":
        parts.append(f"formula={opts.get('formula', '')}")
        refs = opts.get("referencedFieldIds", [])
        if refs:
            parts.append(f"refs={', '.join(refs)}")
        parts.append(f"result={opts.get('result', {}).get('type')}")
    elif opts:
        parts.append(json.dumps(opts, ensure_ascii=False))

    return " | ".join(parts)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def autosize_columns(ws, max_width=80):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def build_rows(schema):
    field_rows = []
    choice_rows = []
    formula_rows = []
    table_rows = []

    for base_key in ("live", "original"):
        base = schema[base_key]
        for table_id, table in base["tables"].items():
            if not table:
                continue

            table_rows.append(
                {
                    "Base": base["name"],
                    "App ID": base["appId"],
                    "Table Name": table["tableName"],
                    "Table ID": table_id,
                    "Primary Field ID": table["primaryFieldId"],
                    "Field Count": table["fieldCount"],
                }
            )

            orig_table = schema["original" if base_key == "live" else "live"]["tables"][table_id]
            orig_by_id = {f["id"]: f for f in orig_table["fields"]}

            for idx, field in enumerate(table["fields"], 1):
                orig_field = orig_by_id.get(field["id"])
                type_diff = ""
                if orig_field and orig_field["type"] != field["type"]:
                    other_base = "Original" if base_key == "live" else "Live"
                    type_diff = f"{other_base} type: {orig_field['type']}"

                field_rows.append(
                    {
                        "Base": base["name"],
                        "App ID": base["appId"],
                        "Table Name": table["tableName"],
                        "Table ID": table_id,
                        "Field #": idx,
                        "Field Name": field["name"],
                        "Field ID": field["id"],
                        "Type": field["type"],
                        "Format / Options": format_field_options(field),
                        "Description": field.get("description", ""),
                        "Type Diff vs Other Base": type_diff,
                    }
                )

                if field["type"] in ("singleSelect", "multipleSelects"):
                    for choice in field.get("options", {}).get("choices", []):
                        choice_rows.append(
                            {
                                "Base": base["name"],
                                "Table Name": table["tableName"],
                                "Table ID": table_id,
                                "Field Name": field["name"],
                                "Field ID": field["id"],
                                "Choice Name": choice.get("name"),
                                "Choice ID": choice.get("id"),
                                "Color": choice.get("color", ""),
                            }
                        )

                if field["type"] == "formula":
                    opts = field.get("options", {})
                    formula_rows.append(
                        {
                            "Base": base["name"],
                            "Table Name": table["tableName"],
                            "Table ID": table_id,
                            "Field Name": field["name"],
                            "Field ID": field["id"],
                            "Formula": opts.get("formula", ""),
                            "Referenced Field IDs": ", ".join(opts.get("referencedFieldIds", [])),
                            "Result Type": opts.get("result", {}).get("type", ""),
                            "Result Options": json.dumps(opts.get("result", {}).get("options", {})),
                        }
                    )

    return table_rows, field_rows, choice_rows, formula_rows


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    style_header_row(ws)
    for row in rows:
        ws.append([row[h] for h in headers])
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def create_workbook(schema):
    table_rows, field_rows, choice_rows, formula_rows = build_rows(schema)

    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme["A1"] = "Airtable Schema Export"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = "Generated from airtable_8_tables_schema.json"
    readme["A4"] = "Upload this file to Google Drive, then Open with > Google Sheets."
    readme["A6"] = "Tabs:"
    readme["A7"] = "Tables — one row per table per base"
    readme["A8"] = "All Fields — full field list for all 8 table instances"
    readme["A9"] = "Select Choices — every single-select choice with choice ID"
    readme["A10"] = "Formulas — formula fields with full expressions"
    readme["A12"] = "Live base app ID: appwocx9mhLR8Mh33"
    readme["A13"] = "Original base app ID: appZoN6xBB9mDv8h4"
    readme.column_dimensions["A"].width = 90

    for title, rows in (
        ("Tables", table_rows),
        ("All Fields", field_rows),
        ("Select Choices", choice_rows),
        ("Formulas", formula_rows),
    ):
        ws = wb.create_sheet(title)
        write_sheet(ws, rows)

    wb.save(XLSX_PATH)


def sheet_name(base_name, table_name):
    label = f"{base_name[:1]}-{table_name}"
    label = re.sub(r"[\[\]\*\?:/\\]", "-", label)
    return label[:31]


def create_apps_script(schema):
    table_rows, field_rows, choice_rows, formula_rows = build_rows(schema)
    payload = {
        "tables": table_rows,
        "fields": field_rows,
        "choices": choice_rows,
        "formulas": formula_rows,
    }
    data_json = json.dumps(payload, ensure_ascii=False)
    data_json = data_json.replace("\\", "\\\\").replace("'", "\\'")

    script = f"""/**
 * Airtable Schema Export → Google Sheets
 *
 * Setup:
 * 1. Create a new Google Sheet (or open an existing one)
 * 2. Extensions → Apps Script
 * 3. Paste this entire file, save
 * 4. Run createSchemaWorkbook once and approve permissions
 */

const SCHEMA_DATA = JSON.parse('{data_json}');

function createSchemaWorkbook() {{
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  writeSheet_(ss, 'Tables', SCHEMA_DATA.tables);
  writeSheet_(ss, 'All Fields', SCHEMA_DATA.fields);
  writeSheet_(ss, 'Select Choices', SCHEMA_DATA.choices);
  writeSheet_(ss, 'Formulas', SCHEMA_DATA.formulas);
  writeReadme_(ss);
  SpreadsheetApp.getUi().alert('Schema workbook created', 'Tabs populated: Tables, All Fields, Select Choices, Formulas, README', SpreadsheetApp.getUi().ButtonSet.OK);
}}

function writeReadme_(ss) {{
  const sheet = getOrCreateSheet_(ss, 'README');
  sheet.clear();
  const lines = [
    ['Airtable Schema Export'],
    [''],
    ['Live base app ID: appwocx9mhLR8Mh33'],
    ['Original base app ID: appZoN6xBB9mDv8h4'],
    [''],
    ['Tabs:'],
    ['Tables — one row per table per base'],
    ['All Fields — full field list for all 8 table instances'],
    ['Select Choices — every single-select choice with choice ID'],
    ['Formulas — formula fields with full expressions'],
  ];
  sheet.getRange(1, 1, lines.length, 1).setValues(lines);
  sheet.getRange(1, 1).setFontWeight('bold').setFontSize(14);
  sheet.setColumnWidth(1, 520);
}}

function writeSheet_(ss, name, rows) {{
  const sheet = getOrCreateSheet_(ss, name);
  sheet.clear();
  if (!rows || rows.length === 0) {{
    return;
  }}
  const headers = Object.keys(rows[0]);
  sheet.getRange(1, 1, 1, headers.length).setValues([headers]);
  sheet.getRange(1, 1, 1, headers.length)
    .setFontWeight('bold')
    .setBackground('#522A10')
    .setFontColor('#FFE180');
  sheet.setFrozenRows(1);

  const values = rows.map(function(row) {{
    return headers.map(function(header) {{ return row[header]; }});
  }});
  sheet.getRange(2, 1, values.length, headers.length).setValues(values);
  for (var i = 1; i <= headers.length; i++) {{
    sheet.autoResizeColumn(i);
  }}
}}

function getOrCreateSheet_(ss, name) {{
  var sheet = ss.getSheetByName(name);
  if (!sheet) {{
    sheet = ss.insertSheet(name);
  }}
  return sheet;
}}
"""
    APPS_SCRIPT_PATH.parent.mkdir(parents=True, exist_ok=True)
    APPS_SCRIPT_PATH.write_text(script, encoding="utf-8")


def main():
    schema = load_schema()
    create_workbook(schema)
    create_apps_script(schema)
    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {APPS_SCRIPT_PATH}")


if __name__ == "__main__":
    main()
