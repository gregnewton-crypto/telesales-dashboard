#!/usr/bin/env python3
"""Build the Fresh Lead Supply Planner workbook.

Produces an .xlsx that is meant to be dropped into Google Drive and opened as a
Google Sheet. Every derived cell is a live formula, so the team can change the
target, the attempt cap or the headcount and watch the required lead volume
move — nothing is a pasted constant.

The attempt yield curve ships pre-loaded with the figures measured from the IE
Telesales base (Mar-Aug 2026, 7,572 leads / 26,115 calls) so the planner works
the moment it is opened. Pasting a fresh call export into 'Paste Calls' and
flipping the source toggle on the Planner recalculates the curve from live data.

Only aggregate counts are baked in; no lead-level records are written.

Usage: python3 tools/build_lead_supply_sheet.py [output.xlsx]
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BROWN = "522A10"
GOLD = "FFE180"
CREAM = "FFF6D8"
PAPER = "FFFCF2"
LINE = "E8D9B0"
INPUT_BG = "FFF3CD"
ANSWER_BG = "DCEDC8"
WARN_BG = "FADBD8"

# Attempt, dials, connects (>20s), sales — lead-level, one sale per lead,
# measured from the IE Telesales base. See tools/ie_lead_flow_model.py.
BASELINE = [
    (1, 7754, 3510, 213),
    (2, 6372, 2580, 183),
    (3, 5134, 1816, 113),
    (4, 3931, 1329, 62),
    (5, 2111, 637, 24),
    (6, 493, 196, 10),
    (7, 144, 58, 8),
    (8, 59, 27, 2),
]

# Observed spread of dials in a real agent-day (172 agent-days with 10+ dials).
CAPACITY_BENCHMARKS = [
    ("Quiet day (25th percentile)", 82, 37, 2.3),
    ("Typical day (median)", 129, 56, 3.3),
    ("Strong day (75th percentile)", 198, 80, 4.4),
    ("Best days (90th percentile)", 295, 100, 4.9),
]

CLOSED_STATUSES = ["Success", "Not interested", "Invalid", "Unqualified"]

# Pasted exports start on row 5. Ranges are bounded rather than whole-column so
# the header and the worked example above them can never land in a COUNTIFS.
LAST = 100000
CALLS_B = f"'Paste Calls'!$B$5:$B${LAST}"
CALLS_D = f"'Paste Calls'!$D$5:$D${LAST}"
CALLS_E = f"'Paste Calls'!$E$5:$E${LAST}"
LEADS_A = f"'Paste Leads'!$A$5:$A${LAST}"
LEADS_B = f"'Paste Leads'!$B$5:$B${LAST}"
LEADS_C = f"'Paste Leads'!$C$5:$C${LAST}"
LEADS_D = f"'Paste Leads'!$D$5:$D${LAST}"

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=16, color=GOLD)
    ws["A1"].fill = PatternFill("solid", fgColor=BROWN)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=10, color="6B5B47")
        ws.merge_cells("A2:H2")


def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color=BROWN)
    c.fill = PatternFill("solid", fgColor=GOLD)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


def header_row(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = Font(bold=True, size=10, color=GOLD)
        c.fill = PatternFill("solid", fgColor=BROWN)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def field(ws, row, label, value, note="", kind="input", fmt=None):
    """One label / value / note line. kind drives the colour coding."""
    ws.cell(row=row, column=1, value=label).font = Font(size=10)
    c = ws.cell(row=row, column=2, value=value)
    c.border = BORDER
    if kind == "input":
        c.fill = PatternFill("solid", fgColor=INPUT_BG)
        c.font = Font(bold=True, size=10)
    elif kind == "answer":
        c.fill = PatternFill("solid", fgColor=ANSWER_BG)
        c.font = Font(bold=True, size=12, color="1B5E20")
    else:
        c.fill = PatternFill("solid", fgColor=PAPER)
        c.font = Font(size=10)
    if fmt:
        c.number_format = fmt
    if note:
        n = ws.cell(row=row, column=3, value=note)
        n.font = Font(size=9, italic=True, color="6B5B47")
    return c


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ───────────────────────────── Start Here ──────────────────────────────────
def build_start(ws):
    title(ws, "Fresh Lead Supply Planner",
          "How many new leads the dialler needs each day to keep the team selling.")
    widths(ws, {"A": 4, "B": 104})

    lines = [
        ("h", "What this answers"),
        ("p", "One number: how many brand-new leads must land in the dialler each "
              "working day. It is derived, not guessed — from how a lead actually "
              "performs on its 1st, 2nd, 3rd and 4th attempt in your own data."),
        ("h", "The idea in one line"),
        ("p", "A lead is not a lead, it is a small budget of attempts. Each attempt "
              "returns less than the one before, so a lead is used up after a few "
              "dials. Divide the team's daily dialling capacity by how many dials a "
              "lead absorbs before it is spent, and you get the number of fresh "
              "leads a day it takes to avoid re-dialling exhausted list."),
        ("h", "Two formulas, and you take the smaller one"),
        ("p", "Target-driven:   fresh leads/day  =  sales target per day  ÷  sales per lead"),
        ("p", "Capacity-driven: fresh leads/day  =  (agents × dials per agent-day)  ÷  "
              "dials per lead"),
        ("p", "Order the LOWER of the two. If the target-driven number is higher, no "
              "amount of lead buying will fix it — the team cannot physically work "
              "them, and surplus leads simply age until they are worth half as much. "
              "That case is a hiring problem, and the Planner flags it."),
        ("h", "How to use it"),
        ("p", "1.  Open the Planner tab. Everything shaded yellow is yours to change; "
              "green cells are the answers."),
        ("p", "2.  Set your sales target, how many agents actually dial on a normal "
              "day, and your attempt cap."),
        ("p", "3.  Read 'ORDER THIS MANY FRESH LEADS PER DAY'."),
        ("p", "4.  Once a month, refresh the yield curve: export calls from Airtable "
              "into Paste Calls, then set the source toggle on the Planner to Auto."),
        ("h", "What counts as a 'fresh' lead"),
        ("p", "Both cliffs matter, and they are separate. In the measured data a lead "
              "holds its value for about 30 days, then falls off sharply; and it holds "
              "its value for about 3 attempts, then falls off sharply. A lead is fresh "
              "while it is under both limits. Attempt 4 and beyond on a lead over 30 "
              "days old converts at 1.12 per 100 dials against 2.70 inside the window."),
        ("h", "Refreshing the data"),
        ("p", "Paste Calls needs one row per call: Lead ID, attempt number, call date, "
              "duration in seconds, outcome. In the IE Airtable base those are "
              "'Databowl LeadId', 'Call # for Lead', 'Date', 'Call Duration' and "
              "'Lead Status' on the Adversus Ireland table."),
        ("p", "Paste Leads needs one row per lead: Lead ID, lead date, times called, "
              "status — 'Databowl Lead ID', 'Lead Date', 'Times Lead has been Called' "
              "and 'Adversus Lead Status (single select)' on the Ireland Leads table."),
        ("h", "Where the shipped numbers come from"),
        ("p", "The baseline yield curve is measured from 7,572 leads and 26,115 calls "
              "in the IE Telesales base between 20 March and 1 September 2026. Sales "
              "are counted once per lead, on the attempt where the sale was stamped. "
              "The curve is survivorship-corrected: it asks what share of the leads "
              "that actually received attempt N converted on attempt N, rather than "
              "the misleading cut of conversion by total attempts, which flatters low "
              "attempt counts because a lead that sells stops being dialled."),
    ]
    r = 4
    for kind, text in lines:
        c = ws.cell(row=r, column=2, value=text)
        if kind == "h":
            c.font = Font(bold=True, size=12, color=BROWN)
            r += 1
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 15 * (len(text) // 95 + 1)
            r += 1
        r += 0
    ws.sheet_view.showGridLines = False


# ─────────────────────────────── Planner ───────────────────────────────────
def build_planner(ws):
    title(ws, "Planner",
          "Yellow = your inputs. Green = the answer. Everything else is calculated.")
    widths(ws, {"A": 46, "B": 14, "C": 62})

    section(ws, 4, "1 · THE TARGET")
    field(ws, 5, "Sales target per working day", 20)
    field(ws, 6, "Working days per week", 5)
    field(ws, 7, "Sales target per week", "=B5*B6", "", "calc", "0.0")

    section(ws, 9, "2 · HOW YOU WORK A LEAD")
    field(ws, 10, "Attempt cap (stop dialling after this many)", 4,
          "Set by economics, not habit — see the break-even below.")
    field(ws, 11, "Retire leads older than (days)", 30,
          "Yield collapses past this point in your data.")
    dv = DataValidation(type="list", formula1='"Baseline,Auto"', allow_blank=False)
    ws.add_data_validation(dv)
    field(ws, 12, "Yield curve source", "Baseline",
          "Baseline = measured Mar-Aug 2026. Auto = recalculated from Paste Calls.")
    dv.add(ws["B12"])
    field(ws, 13, "Dials a lead absorbs before it is spent",
          "=INDEX('Yield Curve'!$H$28:$H$35,$B$10)", "", "calc", "0.00")
    field(ws, 14, "Sales per lead at this cap",
          "=INDEX('Yield Curve'!$I$28:$I$35,$B$10)", "", "calc", "0.0000")
    field(ws, 15, "Sales per 100 dials at this cap", "=IFERROR(100*B14/B13,\"\")",
          "", "calc", "0.00")

    section(ws, 17, "3 · DIALLING CAPACITY")
    field(ws, 18, "Agents dialling on an average day", 2.5,
          "Bodies actually on the phone, not headcount on the payroll.")
    field(ws, 19, "Dials per agent per day", 128,
          "Your measured median. See the Capacity tab before changing this.")
    field(ws, 20, "Total dials available per day", "=B18*B19", "", "calc", "0")

    section(ws, 22, "4 · MONEY")
    field(ws, 23, "Cost per fresh lead", 3.00, "", "input", "#,##0.00")
    field(ws, 24, "Agent cost per hour, fully loaded", 18.00, "", "input", "#,##0.00")
    field(ws, 25, "Paid hours per agent-day", 7.5)
    field(ws, 26, "Cost of one dial", "=IFERROR(B24*B25/B19,\"\")",
          "Agent time only.", "calc", "#,##0.00")
    field(ws, 27, "Gross margin per sale", 150.00, "", "input", "#,##0.00")

    section(ws, 29, "THE ANSWER")
    field(ws, 30, "Leads/day the sales target requires", "=IFERROR(B5/B14,\"\")",
          "", "calc", "0")
    field(ws, 31, "Leads/day the team can actually work", "=IFERROR(B20/B13,\"\")",
          "", "calc", "0")
    c = field(ws, 32, "ORDER THIS MANY FRESH LEADS PER DAY",
              "=IFERROR(MIN(B30,B31),\"\")", "", "answer", "0")
    ws.cell(row=32, column=1).font = Font(bold=True, size=11, color=BROWN)
    field(ws, 33, "…per week", "=IFERROR(B32*B6,\"\")", "", "answer", "0")
    field(ws, 34, "Sales/day that volume delivers", "=IFERROR(B32*B14,\"\")",
          "", "calc", "0.0")
    field(ws, 35, "Gap to target", "=IFERROR(B34-B5,\"\")",
          "Negative means the target is out of reach at this headcount.",
          "calc", "0.0;[Red]-0.0")

    section(ws, 37, "IS THE TARGET EVEN REACHABLE?")
    field(ws, 38, "Agents needed to work the target volume",
          "=IFERROR(B30*B13/B19,\"\")", "", "calc", "0.0")
    field(ws, 39, "Agents you are short",
          "=IFERROR(MAX(0,B38-B18),\"\")", "", "calc", "0.0")
    v = field(ws, 40, "Verdict",
              '=IF(B30>B31,"AGENT-SHORT — buy leads only up to capacity and hire, '
              'or surplus leads will age out","LEAD SUPPLY IS THE ONLY GAP — '
              'buying to target is enough")', "", "calc")
    v.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("B40:C40")
    ws.row_dimensions[40].height = 30

    section(ws, 42, "HOW BIG THE LIVE LIST SHOULD BE")
    field(ws, 43, "Days from first dial to last dial", 9,
          "Your measured median. p75 is 27 days.")
    field(ws, 44, "Workable leads to hold in the dialler",
          "=IFERROR(B32*B43,\"\")", "Little's law: arrival rate × cycle time.",
          "calc", "0")
    field(ws, 45, "Days of untouched cover to keep", 3,
          "Buffer against lumpy supplier delivery.")
    field(ws, 46, "Minimum untouched leads in hand",
          "=IFERROR(B32*B45,\"\")", "", "calc", "0")

    section(ws, 48, "UNIT ECONOMICS")
    field(ws, 49, "Cost per sale", "=IFERROR((B23+B26*B13)/B14,\"\")",
          "Lead cost plus the agent time to work it.", "calc", "#,##0.00")
    field(ws, 50, "Profit per lead", "=IFERROR(B14*B27-B23-B26*B13,\"\")",
          "", "calc", "#,##0.00")
    field(ws, 51, "Profit per day at the ordered volume",
          "=IFERROR(B32*B50,\"\")", "", "calc", "#,##0.00")
    field(ws, 52, "Most you could pay per lead and still break even",
          "=IFERROR(B14*B27-B26*B13,\"\")",
          "Above this price the lead loses money at this cap.", "calc", "#,##0.00")

    section(ws, 54, "WHERE YOU ARE TODAY  (type in your current actuals)")
    field(ws, 55, "Fresh leads arriving per day now", 44)
    field(ws, 56, "Dials per day now", 357)
    field(ws, 57, "Sales per day now", 7)
    field(ws, 58, "Your sales per 100 dials", "=IFERROR(100*B57/B56,\"\")",
          "", "calc", "0.00")
    field(ws, 59, "Supply as a share of what the dialler eats",
          "=IFERROR(B55/B31,\"\")",
          "Under 100% means the shortfall is being filled with old list.",
          "calc", "0%")
    field(ws, 60, "Dials spent per new lead arriving",
          "=IFERROR(B56/B55,\"\")",
          "Compare with row 13. Much higher means you are recycling.",
          "calc", "0.0")
    field(ws, 61, "Sales/day if those same dials hit a fresh list",
          "=IFERROR(B56*B15/100,\"\")", "", "calc", "0.0")
    field(ws, 62, "Prize from fixing supply alone",
          "=IFERROR(B56*B15/100-B57,\"\")",
          "Extra sales per day at today's headcount and today's dial volume.",
          "answer", "0.0")

    for row in (32, 33, 62):
        ws.cell(row=row, column=2).border = Border(
            left=Side(style="medium", color="2E7D32"),
            right=Side(style="medium", color="2E7D32"),
            top=Side(style="medium", color="2E7D32"),
            bottom=Side(style="medium", color="2E7D32"))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


# ───────────────────────────── Yield Curve ─────────────────────────────────
def build_yield(ws):
    title(ws, "Yield Curve",
          "What each successive attempt on a lead is actually worth.")
    widths(ws, {"A": 10, "B": 11, "C": 11, "D": 10, "E": 11,
                "F": 11, "G": 13, "H": 13, "I": 13, "J": 13, "K": 15})

    cols = ["Attempt", "Dials", "Connects", "Sales", "Still in play",
            "Connect rate", "Sales per 100 dials", "Dials per sale"]

    section(ws, 4, "A · RECALCULATED FROM 'PASTE CALLS'  (blank until you paste)")
    header_row(ws, 5, cols)
    for i in range(8):
        r = 6 + i
        n = i + 1
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2,
                value=f"=COUNTIFS({CALLS_B},$A{r})")
        ws.cell(row=r, column=3,
                value=f"=COUNTIFS({CALLS_B},$A{r},{CALLS_D},\">\"&Capacity!$B$26)")
        ws.cell(row=r, column=4,
                value=f"=COUNTIFS({CALLS_B},$A{r},{CALLS_E},Capacity!$B$27)")
        ws.cell(row=r, column=5, value=f"=IFERROR(B{r}/$B$6,\"\")")
        ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/B{r},\"\")")
        ws.cell(row=r, column=7, value=f"=IFERROR(100*D{r}/B{r},\"\")")
        ws.cell(row=r, column=8, value=f"=IFERROR(B{r}/D{r},\"\")")

    section(ws, 15, "B · BASELINE — MEASURED 20 MAR – 1 SEP 2026")
    header_row(ws, 16, cols)
    for i, (n, dials, conn, sales) in enumerate(BASELINE):
        r = 17 + i
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=dials)
        ws.cell(row=r, column=3, value=conn)
        ws.cell(row=r, column=4, value=sales)
        ws.cell(row=r, column=5, value=f"=IFERROR(B{r}/$B$17,\"\")")
        ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/B{r},\"\")")
        ws.cell(row=r, column=7, value=f"=IFERROR(100*D{r}/B{r},\"\")")
        ws.cell(row=r, column=8, value=f"=IFERROR(B{r}/D{r},\"\")")

    for r in list(range(6, 14)) + list(range(17, 25)):
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            if col in (5, 6):
                c.number_format = "0.0%"
            elif col == 7:
                c.number_format = "0.00"
            elif col == 8:
                c.number_format = "0"
            else:
                c.number_format = "#,##0"

    section(ws, 26, "C · CUMULATIVE BY ATTEMPT CAP — THIS IS WHAT THE PLANNER READS")
    header_row(ws, 27, [
        "Cap at", "Dials/lead\n(auto)", "Sales/lead\n(auto)", "Per 100\n(auto)",
        "Dials/lead\n(baseline)", "Sales/lead\n(baseline)", "Per 100\n(baseline)",
        "Dials/lead\nIN USE", "Sales/lead\nIN USE",
        "Break-even lead price\n(× cost of one dial)"])
    for i in range(8):
        r = 28 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2,
                value=f"=IFERROR(SUMIF($A$6:$A$13,\"<=\"&$A{r},$B$6:$B$13)/$B$6,\"\")")
        ws.cell(row=r, column=3,
                value=f"=IFERROR(SUMIF($A$6:$A$13,\"<=\"&$A{r},$D$6:$D$13)/$B$6,\"\")")
        ws.cell(row=r, column=4, value=f"=IFERROR(100*C{r}/B{r},\"\")")
        ws.cell(row=r, column=5,
                value=f"=IFERROR(SUMIF($A$17:$A$24,\"<=\"&$A{r},$B$17:$B$24)/$B$17,\"\")")
        ws.cell(row=r, column=6,
                value=f"=IFERROR(SUMIF($A$17:$A$24,\"<=\"&$A{r},$D$17:$D$24)/$B$17,\"\")")
        ws.cell(row=r, column=7, value=f"=IFERROR(100*F{r}/E{r},\"\")")
        ws.cell(row=r, column=8, value=f"=IF(Planner!$B$12=\"Auto\",B{r},E{r})")
        ws.cell(row=r, column=9, value=f"=IF(Planner!$B$12=\"Auto\",C{r},F{r})")
        if i == 0:
            ws.cell(row=r, column=10, value="n/a")
        else:
            p = r - 1
            ws.cell(row=r, column=10,
                    value=f"=IFERROR((H{r}-H{p})*I{p}/(I{r}-I{p})-H{p},\"\")")
        for col in range(1, 11):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            if col in (2, 5, 8):
                c.number_format = "0.00"
            elif col in (3, 6, 9):
                c.number_format = "0.0000"
            elif col in (4, 7, 10):
                c.number_format = "0.00"
        for col in (8, 9):
            ws.cell(row=r, column=col).font = Font(bold=True, color=BROWN)

    notes = [
        "Read column K like this: raising the cap from 3 to 4 is worth doing as long "
        "as a fresh lead costs more than 1.7× what one dial costs in agent time.",
        "Tightening the cap only pays when fresh leads are cheap enough to replace "
        "the attempts you give up. If agents are sitting at part occupancy, a spare "
        "dial is nearly free and the cap is not your problem — supply is.",
        "'Still in play' is the share of leads that survive to receive that attempt. "
        "It falls because leads sell, refuse, or turn out to be junk, not only "
        "because the team stops dialling.",
        "Treat attempts 6 and beyond with suspicion. Those rows look strong, but "
        "almost nothing reaches them by policy — they are leads an agent chose to "
        "keep chasing, usually a booked callback. That is a hand-picked group, not "
        "evidence that a 7th cold dial pays. Do not plan above a cap of 5 on it.",
    ]
    r = 38
    for n in notes:
        c = ws.cell(row=r, column=1, value=n)
        c.font = Font(size=9, italic=True, color="6B5B47")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.row_dimensions[r].height = 28
        r += 1
    ws.sheet_view.showGridLines = False


# ────────────────────────────── Scenarios ──────────────────────────────────
def build_scenarios(ws):
    title(ws, "Scenarios",
          "Fresh leads per day, and the headcount it takes to work them.")
    widths(ws, {"A": 24, **{get_column_letter(i): 13 for i in range(2, 8)}})

    caps = [2, 3, 4, 5, 6]
    targets = [5, 10, 15, 20, 25, 30, 40, 50]

    section(ws, 4, "FRESH LEADS NEEDED PER DAY")
    ws.cell(row=5, column=1, value="Sales/day target")
    header_row(ws, 5, [f"cap at {c}" for c in caps], start_col=2)
    ws.cell(row=5, column=1).font = Font(bold=True, size=10, color=GOLD)
    ws.cell(row=5, column=1).fill = PatternFill("solid", fgColor=BROWN)
    ws.cell(row=5, column=1).border = BORDER
    for i, t in enumerate(targets):
        r = 6 + i
        ws.cell(row=r, column=1, value=t).font = Font(bold=True)
        for j, cap in enumerate(caps):
            c = ws.cell(row=r, column=2 + j,
                        value=f"=IFERROR($A{r}/INDEX('Yield Curve'!$I$28:$I$35,{cap}),\"\")")
            c.number_format = "#,##0"
            c.border = BORDER

    section(ws, 16, "AGENTS ON THE PHONE NEEDED TO WORK THAT VOLUME")
    ws.cell(row=17, column=1, value="Sales/day target")
    header_row(ws, 17, [f"cap at {c}" for c in caps], start_col=2)
    ws.cell(row=17, column=1).font = Font(bold=True, size=10, color=GOLD)
    ws.cell(row=17, column=1).fill = PatternFill("solid", fgColor=BROWN)
    ws.cell(row=17, column=1).border = BORDER
    for i, t in enumerate(targets):
        r = 18 + i
        ws.cell(row=r, column=1, value=t).font = Font(bold=True)
        for j, cap in enumerate(caps):
            c = ws.cell(
                row=r, column=2 + j,
                value=f"=IFERROR($A{r}/INDEX('Yield Curve'!$I$28:$I$35,{cap})"
                      f"*INDEX('Yield Curve'!$H$28:$H$35,{cap})/Planner!$B$19,\"\")")
            c.number_format = "0.0"
            c.border = BORDER

    section(ws, 28, "COST PER SALE AT EACH CAP")
    ws.cell(row=29, column=1, value="Cost per fresh lead")
    header_row(ws, 29, [f"cap at {c}" for c in caps], start_col=2)
    ws.cell(row=29, column=1).font = Font(bold=True, size=10, color=GOLD)
    ws.cell(row=29, column=1).fill = PatternFill("solid", fgColor=BROWN)
    ws.cell(row=29, column=1).border = BORDER
    for i, price in enumerate([1, 2, 3, 5, 8, 12]):
        r = 30 + i
        c0 = ws.cell(row=r, column=1, value=price)
        c0.font = Font(bold=True)
        c0.number_format = "#,##0.00"
        for j, cap in enumerate(caps):
            c = ws.cell(
                row=r, column=2 + j,
                value=f"=IFERROR(($A{r}+Planner!$B$26"
                      f"*INDEX('Yield Curve'!$H$28:$H$35,{cap}))"
                      f"/INDEX('Yield Curve'!$I$28:$I$35,{cap}),\"\")")
            c.number_format = "#,##0"
            c.border = BORDER

    note = ("The lowest cost per sale in each row of the bottom grid is the cap you "
            "should be running at that lead price. Cheap leads reward a tight cap and "
            "high volume; expensive leads reward working each lead harder.")
    c = ws.cell(row=37, column=1, value=note)
    c.font = Font(size=9, italic=True, color="6B5B47")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A37:G37")
    ws.row_dimensions[37].height = 30
    ws.sheet_view.showGridLines = False


# ─────────────────────────────── Capacity ──────────────────────────────────
def build_capacity(ws):
    title(ws, "Capacity",
          "Is the dials-per-agent-day number on the Planner physically possible?")
    widths(ws, {"A": 42, "B": 14, "C": 14, "D": 14, "E": 60})

    section(ws, 4, "WHAT A REAL AGENT-DAY LOOKED LIKE  (172 agent-days, Mar-Aug 2026)")
    header_row(ws, 5, ["", "Dials", "Connects", "Talk hours"])
    for i, (label, d, c, t) in enumerate(CAPACITY_BENCHMARKS):
        r = 6 + i
        ws.cell(row=r, column=1, value=label).font = Font(size=10)
        for col, val, fmt in ((2, d, "#,##0"), (3, c, "#,##0"), (4, t, "0.0")):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.border = BORDER

    section(ws, 11, "SANITY-CHECK YOUR PLANNING ASSUMPTION")
    field(ws, 12, "Dials per agent-day assumed on the Planner",
          "=Planner!$B$19", "", "calc", "#,##0")
    field(ws, 13, "Connect rate you expect", 0.40,
          "Blended across attempts 1 to the cap.", "input", "0%")
    field(ws, 14, "Average talk time per connect (minutes)", 3.5)
    field(ws, 15, "Seconds of wrap and dialling per dial", 25)
    field(ws, 16, "Talk hours implied", "=IFERROR(B12*B13*B14/60,\"\")",
          "", "calc", "0.0")
    field(ws, 17, "Wrap and dial hours implied", "=IFERROR(B12*B15/3600,\"\")",
          "", "calc", "0.0")
    field(ws, 18, "Total occupied hours", "=IFERROR(B16+B17,\"\")", "", "calc", "0.0")
    field(ws, 19, "Paid hours per agent-day", "=Planner!$B$25", "", "calc", "0.0")
    field(ws, 20, "Occupancy", "=IFERROR(B18/B19,\"\")", "", "calc", "0%")
    v = field(ws, 21, "Verdict",
              '=IF(B20>0.95,"IMPOSSIBLE — lower the dials assumption",'
              'IF(B20>0.85,"VERY TIGHT — no room for breaks or admin",'
              'IF(B20<0.5,"LOTS OF SLACK — the phones are not the constraint",'
              '"REALISTIC")))', "", "calc")
    v.alignment = Alignment(wrap_text=True, vertical="center")

    section(ws, 23, "SETTINGS USED ELSEWHERE")
    field(ws, 24, "Measured occupancy, Mar-Aug 2026", 0.44,
          "3.3 talk hours against a 7.5 hour shift.", "calc", "0%")
    field(ws, 26, "A call counts as connected above (seconds)", 20,
          "12.3% of calls last exactly 20s — that is the dialler's ring timeout.")
    field(ws, 27, "Outcome value that means a sale", "Success",
          "Must match the text in your export exactly.")

    note = ("Occupancy below about 60% says the team is starved of list, not short of "
            "hours: they have time on the clock and nothing worth dialling. That is "
            "the state the measured data is in, and it is why buying more leads works "
            "before hiring does.")
    c = ws.cell(row=29, column=1, value=note)
    c.font = Font(size=9, italic=True, color="6B5B47")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A29:E29")
    ws.row_dimensions[29].height = 30
    ws.sheet_view.showGridLines = False


# ────────────────────────────── List Health ────────────────────────────────
def build_health(ws):
    """Daily health check driven entirely by the Paste Leads tab."""
    title(ws, "List Health",
          "How much genuinely workable list is left, and how long it lasts.")
    widths(ws, {"A": 46, "B": 14, "C": 66})

    closed_terms = "".join(
        f",{LEADS_D},\"<>{s}\"" for s in CLOSED_STATUSES)

    section(ws, 4, "THE LIST RIGHT NOW  (from Paste Leads)")
    field(ws, 5, "Today", "=TODAY()", "", "calc", "yyyy-mm-dd")
    field(ws, 6, "Leads loaded", f"=COUNTA({LEADS_A})", "", "calc", "#,##0")
    field(ws, 7, "Arrived inside the freshness window",
          f"=COUNTIFS({LEADS_B},\">=\"&(TODAY()-Planner!$B$11))",
          "Lead date within the Planner's retirement age.", "calc", "#,##0")
    field(ws, 8, "…of those, still open and under the attempt cap",
          f"=COUNTIFS({LEADS_B},\">=\"&(TODAY()-Planner!$B$11),"
          f"{LEADS_C},\"<\"&Planner!$B$10" + closed_terms + ")",
          "This is your real workable stock.", "answer", "#,##0")
    field(ws, 9, "Open but already at the attempt cap",
          f"=COUNTIFS({LEADS_B},\">=\"&(TODAY()-Planner!$B$11),"
          f"{LEADS_C},\">=\"&Planner!$B$10" + closed_terms + ")",
          "Burnt inside the window — the symptom of running short on fresh list.",
          "calc", "#,##0")
    field(ws, 10, "Open but past the retirement age",
          f"=COUNTIFS({LEADS_B},\"<\"&(TODAY()-Planner!$B$11)"
          + closed_terms + ")",
          "Worth roughly half. Archive rather than dial.", "calc", "#,##0")

    section(ws, 12, "RUNWAY")
    field(ws, 13, "Workable stock", "=B8", "", "calc", "#,##0")
    field(ws, 14, "Leads consumed per day at plan", "=Planner!$B$32", "", "calc", "0")
    field(ws, 15, "Days of cover left", "=IFERROR(B13/B14,\"\")", "", "answer", "0.0")
    field(ws, 16, "Target days of cover", "=Planner!$B$45", "", "calc", "0")
    v = field(ws, 17, "Verdict",
              '=IF(B15<1,"EMPTY — agents are re-dialling exhausted list today",'
              'IF(B15<B16,"BELOW BUFFER — order now",'
              'IF(B15>B16*4,"OVERSTOCKED — leads will age out before they are worked",'
              '"HEALTHY")))', "", "calc")
    v.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("B17:C17")

    section(ws, 19, "AGEING PROFILE")
    header_row(ws, 20, ["Lead age", "Open leads", "What to do with them"])
    bands = [
        (0, 7, "Prime. Should be on attempt 1 or 2."),
        (8, 14, "Still strong. Attempts 2 to 3."),
        (15, 30, "Last of the value. Finish the cap here."),
        (31, 60, "Yield roughly halves. Stop unless nothing else is loaded."),
        (61, 3650, "Effectively dead. 0.59 sales per 100 dials in the measured data."),
    ]
    for i, (lo, hi, advice) in enumerate(bands):
        r = 21 + i
        ws.cell(row=r, column=1,
                value=(f"{lo}–{hi} days" if hi < 3650 else f"{lo}+ days"))
        ws.cell(row=r, column=2,
                value=f"=COUNTIFS({LEADS_B},\"<=\"&(TODAY()-{lo}),"
                      f"{LEADS_B},\">=\"&(TODAY()-{hi})"
                      + closed_terms + ")").number_format = "#,##0"
        ws.cell(row=r, column=3, value=advice).font = Font(size=9, color="6B5B47")
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
    ws.sheet_view.showGridLines = False


# ──────────────────────────── Paste targets ────────────────────────────────
def build_paste(ws, headers, blurb, examples):
    """Paste target. Row 4 is the header, data starts at row 5.

    The worked example sits off to the right rather than in the data rows, so a
    user who forgets to delete it cannot skew the yield curve.
    """
    title(ws, ws.title, blurb + "  Paste your export into A5.")
    header_row(ws, 4, [h for h, _ in headers])
    for i, (_, w) in enumerate(headers):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    off = len(headers) + 2
    c = ws.cell(row=3, column=off, value="what it should look like")
    c.font = Font(size=9, bold=True, italic=True, color="6B5B47")
    for i, (h, w) in enumerate(headers):
        col = off + i
        ws.column_dimensions[get_column_letter(col)].width = w
        e = ws.cell(row=4, column=col, value=h)
        e.font = Font(size=9, bold=True, color="9C8B6E")
        e.border = BORDER
    for j, row in enumerate(examples):
        for i, val in enumerate(row):
            e = ws.cell(row=5 + j, column=off + i, value=val)
            e.font = Font(size=9, italic=True, color="9C8B6E")
            e.border = BORDER
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "fresh-lead-supply-planner.xlsx"
    wb = Workbook()

    build_start(wb.active)
    wb.active.title = "Start Here"
    build_planner(wb.create_sheet("Planner"))
    build_scenarios(wb.create_sheet("Scenarios"))
    build_yield(wb.create_sheet("Yield Curve"))
    build_capacity(wb.create_sheet("Capacity"))
    build_health(wb.create_sheet("List Health"))

    build_paste(
        wb.create_sheet("Paste Calls"),
        [("Lead ID", 16), ("Attempt no.", 13), ("Call date", 14),
         ("Duration (s)", 13), ("Outcome", 22)],
        "One row per call. Airtable › Adversus Ireland › Databowl LeadId, "
        "Call # for Lead, Date, Call Duration, Lead Status.",
        [[1234567, 1, "2026-08-03", 14, "Automatic redial"],
         [1234567, 2, "2026-08-06", 812, "Success"],
         [1234568, 1, "2026-08-03", 20, "Automatic redial"]])

    build_paste(
        wb.create_sheet("Paste Leads"),
        [("Lead ID", 16), ("Lead date", 14), ("Times called", 14), ("Status", 26)],
        "One row per lead. Airtable › Ireland Leads › Databowl Lead ID, Lead Date, "
        "Times Lead has been Called, Adversus Lead Status (single select).",
        [[1234567, "2026-08-03", 2, "Success"],
         [1234568, "2026-08-03", 4, "Automatic redial"],
         [1234569, "2026-08-04", 1, "Private callback"]])

    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
